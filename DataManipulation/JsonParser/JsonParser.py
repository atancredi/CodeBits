from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import json

def xlsx_to_json(path):
    wb = load_workbook(filename=path)
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        my_list.append(my_dict)

    data = json.dumps(my_list, sort_keys=True, indent=4, default=str, ensure_ascii=False)
    with open(path.replace(".","_")+"_converted.json", 'w', encoding='utf-8') as f:
        f.write(data)


files = [f for f in os.listdir('.') if os.path.isfile(f)]
for f in files:
    if(".xlsx" in f):
        xlsx_to_json(f)
